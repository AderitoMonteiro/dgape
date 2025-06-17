from django.shortcuts import render,get_object_or_404
from .models import sala,inventario_mobiliario_eleitoral,inventario_equipamento_eleitoral, mobiliario_eleitoral,equipamento_eleitoral,inventario_mobiliario,equipamento,mobiliario,inventario_equipamento
from kit_eleitoral.models import conselho
from django.http import HttpResponse,JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Font
from django.core.paginator import Paginator
from django.core.serializers import serialize
from datetime import datetime
from django.db import connection
import openpyxl


def inventario_equipamento_home(request):

    equipamento_list = equipamento.objects.all().filter(status=1)
    equipamento_inventario_list = inventario_equipamento.objects.all().filter(status=1)

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    kec.descricao as conselho,
                    die.obs,
                    die.provinencia,
                    de.descricao,
                    de.modelo,
                    IFNULL(de.marca,'') as marca,
                    de.mac_address,
                    de.serial_number,
                    IFNULL(sa.descricao,'') as sala
                    FROM 
                    departamentos_inventario_equipamento as die
                    left join departamentos_equipamento as de on die.equipamento_id=de.id
                    left join kit_eleitoral_conselho as kec on de.conselho=kec.id  
                    left join departamentos_sala as sa on de.sala=sa.id                  
                    where die.status=1
                '''
    with connection.cursor() as cursor:
          cursor.execute(query)


          colunas = [col[0] for col in cursor.description] 
          resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]

          paginator = Paginator(resultados, 7)
          page_number = request.GET.get("page")  
          paginator_equipamento_inventario = paginator.get_page(page_number)


    return render(request, 'Inventario_equipamento/index.html',{"equipamento":equipamento_list,"equipamento_inventario":paginator_equipamento_inventario})

def inventario_equipamento_eleitoral_home(request):

    equipamento_list = equipamento_eleitoral.objects.all().filter(status=1)
    equipamento_inventario_list = inventario_equipamento.objects.all().filter(status=1)

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    die.localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao,
                    de.modelo,
                    de.marca,
                    de.mac_address,
                    de.serial_number
                    FROM 
                    departamentos_inventario_equipamento_eleitoral as die
                    inner join departamentos_equipamento_eleitoral as de on die.equipamento_id=de.id
                    where die.status=1
                '''
    with connection.cursor() as cursor:
          cursor.execute(query)


          colunas = [col[0] for col in cursor.description] 
          resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]

          paginator = Paginator(resultados, 7)
          page_number = request.GET.get("page")  
          paginator_equipamento_inventario = paginator.get_page(page_number)


    return render(request, 'inventario_equipamento_eleitoral/index.html',{"equipamento":equipamento_list,"equipamento_inventario":paginator_equipamento_inventario})
def inventario_mobiliario_home(request):

    mobiliario_list = mobiliario.objects.all().filter(status=1)
    equipamento_inventario_list = inventario_equipamento.objects.all().filter(status=1)

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    kec.descricao as localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao,
                    ds.descricao as sala
                    FROM 
                    departamentos_inventario_mobiliario as die
                    inner join departamentos_mobiliario as de on die.mobiliario_id=de.id
                    left join kit_eleitoral_conselho as kec on de.conselho=kec.id 
                    left join departamentos_sala as ds on de.sala=ds.id
                    where die.status=1
                '''
    with connection.cursor() as cursor:
          cursor.execute(query)

          colunas = [col[0] for col in cursor.description] 
          resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]

          paginator = Paginator(resultados, 7)
          page_number = request.GET.get("page")  
          paginator_equipamento_mobiliario = paginator.get_page(page_number)


    return render(request, 'Inventario_mobiliario/index.html',{"mobiliario":mobiliario_list,"mobiliario_inventario":paginator_equipamento_mobiliario})
def inventario_mobiliario_eleitoral_home(request):

    mobiliario_list = mobiliario_eleitoral.objects.all().filter(status=1)
    equipamento_inventario_list = inventario_equipamento.objects.all().filter(status=1)

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    die.localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao
                    FROM 
                    departamentos_inventario_mobiliario_eleitoral as die
                    inner join departamentos_mobiliario_eleitoral as de on die.mobiliario_id=de.id
                    where die.status=1
                '''
    with connection.cursor() as cursor:
          cursor.execute(query)

          colunas = [col[0] for col in cursor.description] 
          resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]

          paginator = Paginator(resultados, 7)
          page_number = request.GET.get("page")  
          paginator_equipamento_mobiliario = paginator.get_page(page_number)


    return render(request, 'Inventario_mobiliario_eleitoral/index.html',{"mobiliario":mobiliario_list,"mobiliario_inventario":paginator_equipamento_mobiliario})
    
def index(request):

                    componente ='equipamento'
                     
                    return render(request, 'departamento/index.html',{'componente':componente})

def gestao_equipamento(request):

                      conselho_lis = conselho.objects.all().filter(status=1)
                      sala_lis = sala.objects.all().filter(status=1)  
                      equipamento_list = equipamento.objects.all().filter(status=1)
                      componente ='equipamento'
                      sidebar=request.GET.get('modulo')

                      if sidebar=='gestao':

                          query = '''
                                SELECT 
                                departamentos_equipamento.id,
                                departamentos_equipamento.descricao,
                                mac_address,
                                data_entrada,
                                provinencia,
                                marca,
                                modelo,
                                obs,
                                serial_number,
                                tipo,
                                IFNULL(departamentos_sala.descricao,'') as sala,
                                kit_eleitoral_conselho.descricao as conselho,
                                'get_equipamento_gestao(this)' as sidebar,
                                'sidebar_gestao' as sidebar_descricao
                                FROM 
                                departamentos_equipamento
                                left join departamentos_sala on departamentos_equipamento.sala=departamentos_sala.id
                                left join kit_eleitoral_conselho on departamentos_equipamento.conselho=kit_eleitoral_conselho.id
                                where departamentos_equipamento.status=1 order by departamentos_equipamento.id desc
                                '''
                      else:
                          
                          query = '''
                                SELECT 
                                departamentos_equipamento.id,
                                departamentos_equipamento.descricao,
                                mac_address,
                                data_entrada,
                                provinencia,
                                marca,
                                modelo,
                                obs,
                                serial_number,
                                tipo,
                                IFNULL(departamentos_sala.descricao,'') as sala,
                                kit_eleitoral_conselho.descricao as conselho,
                                'get_equipamento(button)' as sidebar,
                                'sidebar_lancamento' as sidebar_descricao
                                FROM 
                                departamentos_equipamento
                                left join departamentos_sala on departamentos_equipamento.sala=departamentos_sala.id
                                left join kit_eleitoral_conselho on departamentos_equipamento.conselho=kit_eleitoral_conselho.id
                                where departamentos_equipamento.status=1 order by departamentos_equipamento.id desc
                                '''

                      with connection.cursor() as cursor:
                          cursor.execute(query)

                          colunas = [col[0] for col in cursor.description] 
                          equipamento_list = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                          paginator = Paginator(equipamento_list, 7)
                          page_number = request.GET.get("page",1)  
                          paginator_equipamento = paginator.get_page(page_number)

                          if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                             return render(request, 'Equipamento/index.html',{"equipamento":paginator_equipamento,"conselho":conselho_lis,"sala":sala_lis,"componente":componente})
                  
                      return render(request, 'Equipamento/index.html',{"equipamento":paginator_equipamento,"conselho":conselho_lis,"sala":sala_lis,"componente":componente})

def gestao_equipamento_eleitoral(request):

                      

                      equipamento_list = equipamento_eleitoral.objects.all().filter(status=1)
                      paginator = Paginator(equipamento_list, 7)
                      page_number = request.GET.get("page")  
                      paginator_equipamento = paginator.get_page(page_number)
                  
                      return render(request, 'Equipamento_eleitoral/index.html',{"equipamento":paginator_equipamento})
@csrf_exempt
def add_equipamento(request):
  
  
  if request.method == "POST":
            try:
                    descricao= request.POST.get("descricao")
                    data_entrada= request.POST.get("data_entrada")
                    obs= request.POST.get("obs")
                    marca= request.POST.get("marca")
                    provinencia= request.POST.get("provinencia")
                    modelo= request.POST.get("modelo")
                    tipo_item= request.POST.get("tipo_item")
                    sala_id= request.POST.get("sala_id")
                    conselho= request.POST.get("conselho")
                    serial_number= request.POST.get("serial_number")
                    mac_address= request.POST.get("mac_address")
                    user_create= request.POST.get("user_create")

                    validate=equipamento.objects.filter(serial_number=serial_number).count()

                    if validate==0:
                       
                            if tipo_item =="Portatel":
                                            if conselho!="23":
                                                          if descricao !="" and marca !="" and modelo !="" and data_entrada !="" and conselho!=""and tipo_item!="" and serial_number!="" and mac_address !="":

                                                                    equipamento.objects.create(
                                                                                                descricao=descricao,
                                                                                                data_entrada=data_entrada,
                                                                                                obs=obs,
                                                                                                marca=marca,
                                                                                                modelo=modelo,
                                                                                                provinencia=provinencia,
                                                                                                conselho=conselho,
                                                                                                tipo=tipo_item,
                                                                                                serial_number=serial_number,
                                                                                                mac_address=mac_address,
                                                                                                user_create=user_create
                                                                                                  )
                                                                    message='Equipamento registado com sucesso!!'
                                                                    status= 'success'
                                                                    return JsonResponse({'status':status, 'message': message })

                                                          else:
                                                                      message='Erro, tem que preencher todos os campos obrigatorios!!'
                                                                      status= 'error'
                                                                      return JsonResponse({'status':status, 'message': message })
                                            else:
                                                                if descricao !="" and marca !="" and modelo !="" and data_entrada !="" and conselho!="" and sala_id!="" and tipo_item!="" and serial_number!="" and mac_address !="":

                                                                                  equipamento.objects.create(
                                                                                                              descricao=descricao,
                                                                                                              data_entrada=data_entrada,
                                                                                                              obs=obs,
                                                                                                              marca=marca,
                                                                                                              modelo=modelo,
                                                                                                              conselho=conselho,
                                                                                                              tipo=tipo_item,
                                                                                                              provinencia=provinencia,
                                                                                                              sala=sala_id,
                                                                                                              serial_number=serial_number,
                                                                                                              mac_address=mac_address,
                                                                                                              user_create=user_create
                                                                                                                )
                                                                                  message='Equipamento registado com sucesso!!'
                                                                                  status= 'success'
                                                                                  return JsonResponse({'status':status, 'message': message })

                                                                else:
                                                                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                                                                        status= 'error'
                                                                        return JsonResponse({'status':status, 'message': message })

                                                
                                        
                            else: 

                                if conselho!="23":
                                                          if descricao !="" and marca !="" and modelo !="" and data_entrada !="" and conselho!=""and tipo_item!="" and serial_number!="":

                                                                    equipamento.objects.create(
                                                                                                descricao=descricao,
                                                                                                data_entrada=data_entrada,
                                                                                                obs=obs,
                                                                                                marca=marca,
                                                                                                modelo=modelo,
                                                                                                provinencia=provinencia,
                                                                                                conselho=conselho,
                                                                                                tipo=tipo_item,
                                                                                                serial_number=serial_number,
                                                                                                mac_address=mac_address,
                                                                                                user_create=user_create
                                                                                                  )
                                                                    message='Equipamento registado com sucesso!!'
                                                                    status= 'success'
                                                                    return JsonResponse({'status':status, 'message': message })

                                                          else:
                                                            message='Erro, tem que preencher todos os campos obrigatorios!!'
                                                            status= 'error'
                                                            return JsonResponse({'status':status, 'message': message })
                                else:
                                                    if descricao !="" and marca !="" and modelo !="" and data_entrada !="" and conselho!="" and sala_id!="" and tipo_item!="" and serial_number!="":

                                                                      equipamento.objects.create(
                                                                                                  descricao=descricao,
                                                                                                  data_entrada=data_entrada,
                                                                                                  obs=obs,
                                                                                                  marca=marca,
                                                                                                  modelo=modelo,
                                                                                                  conselho=conselho,
                                                                                                  tipo=tipo_item,
                                                                                                  provinencia=provinencia,
                                                                                                  sala=sala_id,
                                                                                                  serial_number=serial_number,
                                                                                                  mac_address=mac_address,
                                                                                                  user_create=user_create
                                                                                                    )
                                                                      message='Equipamento registado com sucesso!!'
                                                                      status= 'success'
                                                                      return JsonResponse({'status':status, 'message': message })

                                                    else:
                                                          message='Erro, tem que preencher todos os campos obrigatorios!!'
                                                          status= 'error'
                                                          return JsonResponse({'status':status, 'message': message })

                                        
                    else:
                             message='Erro, este equipamento já foi registado!!'
                             status= 'error'
                             return JsonResponse({'status':status, 'message': message })

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_equipamento_eleitoral(request):
  
  if request.method == "POST":
            try:
                    data_entrada= request.POST.get("data_entrada")
                    obs= request.POST.get("obs")
                    descricao= request.POST.get("descricao")
                    marca= request.POST.get("marca")
                    modelo= request.POST.get("modelo")
                    serial_number= request.POST.get("serial_number")
                    mac_address= request.POST.get("mac_address")
                    user_create= request.POST.get("user_create")

                    validate=equipamento_eleitoral.objects.filter(descricao=descricao,marca=marca,modelo=modelo,serial_number=serial_number).count()

                    if validate==0:
                          if descricao !="" and marca !="" and modelo !="" and serial_number !="":

                                    equipamento_eleitoral.objects.create(
                                                                descricao=descricao,
                                                                data_entrada=data_entrada,
                                                                obs=obs,
                                                                marca=marca,
                                                                modelo=modelo,
                                                                serial_number=serial_number,
                                                                mac_address=mac_address,
                                                                user_create=user_create
                                                                  )
                                    message='Equipamento registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                          else:
                            message='Erro, tem que preencher todos os campos obrigatorios!!'
                            status= 'error'
                            return JsonResponse({'status':status, 'message': message })
                    else:
                      message='Erro, este equipamento já está registado!!'
                      status= 'error'
                      return JsonResponse({'status':status, 'message': message })

                    

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_equipamento(request):

    if request.method == "POST":
      try:
        
                      equipamento_id = request.POST.get("equipamento_id")

                      query = ''' select 
                                  departamentos_equipamento.id,
                                  departamentos_equipamento.descricao,
                                  mac_address,
                                  data_entrada,
                                  marca,
                                  modelo,
                                  provinencia,
                                  obs,
                                  serial_number,
                                  tipo,
                                  IFNULL(departamentos_sala.descricao,'') as sala,
                                  departamentos_sala.id as sala_id,
                                  kit_eleitoral_conselho.descricao as descricao_conselho,
                                  kit_eleitoral_conselho.id as conselho_id
                                  from 
                                  departamentos_equipamento
                                  left join departamentos_sala on departamentos_equipamento.sala=departamentos_sala.id
                                  left join kit_eleitoral_conselho on departamentos_equipamento.conselho=kit_eleitoral_conselho.id
                                  where departamentos_equipamento.id=%s'''
                      with connection.cursor() as cursor:
                          cursor.execute(query,[equipamento_id])
                          
                          colunas = [col[0] for col in cursor.description] 
                          equipament = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                          print(equipament)
                     
                          return JsonResponse({'resultado': equipament})

      except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_equipamento_eleitoral(request):

    if request.method == "POST":
      try:
                      equipamento_id = request.POST.get("equipamento_id")
                      equipament= equipamento_eleitoral.objects.filter(id=equipamento_id)
                     
                      return JsonResponse(serialize("json", equipament),safe=False)

      except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def editar_equipamento(request):
  
  if request.method == "POST":
            try:
                    equipamento_id = request.POST.get("equipamento_id")
                    obs= request.POST.get("obs")
                    serial_number= request.POST.get("serial_number")
                    conselho= request.POST.get("conselho_edit")
                    sala_id= request.POST.get("sala_id")
                    mac_address= request.POST.get("mac_address")
                    user_update= request.POST.get("user_create")

                    if conselho=="23":

                            if sala_id!="":

                                  equipamento_ob=get_object_or_404(equipamento,id=equipamento_id)
                                  equipamento_ob.obs=obs
                                  equipamento_ob.sala=sala_id
                                  equipamento_ob.conselho=conselho
                                  equipamento_ob.user_update=user_update
                                  equipamento_ob.dateupdate=datetime.now()
                                  equipamento_ob.save()
                                                                                
                                  message='Equipamento alterado com sucesso!!'
                                  status= 'success'
                                  return JsonResponse({'status':status, 'message': message })
                            else:
                              message='Erro, tem que preencher os campos abrigatorio!!'
                              status= 'error'
                              return JsonResponse({'status':status, 'message': message })

                    else:
                          
                          if conselho!="":

                              equipamento_ob=get_object_or_404(equipamento,id=equipamento_id)
                              equipamento_ob.obs=obs
                              equipamento_ob.conselho=conselho
                              equipamento_ob.sala=""
                              equipamento_ob.user_update=user_update
                              equipamento_ob.dateupdate=datetime.now()
                              equipamento_ob.save()
                                                                            
                              message='Equipamento alterado com sucesso!!'
                              status= 'success'
                              return JsonResponse({'status':status, 'message': message })

                          else:
                              message='Erro, tem que preencher os campos abrigatorio!!'
                              status= 'error'
                              return JsonResponse({'status':status, 'message': message })




            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def editar_equipamento_eleitoral(request):
  
  if request.method == "POST":
            try:
                    equipamento_id = request.POST.get("equipamento_id")
                    obs_edit= request.POST.get("obs_edit")
                    user_update= request.POST.get("user_create")

                    equipamento_ob=get_object_or_404(equipamento_eleitoral,id=equipamento_id)
                    equipamento_ob.obs=obs_edit
                    equipamento_ob.user_update=user_update
                    equipamento_ob.dateupdate=datetime.now()
                    equipamento_ob.save()
                                                                  
                    message='Equipamento alterado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })

         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_equipamento(request):
  
  if request.method == "POST":
            try:
                    equipamento_id = request.POST.get("equipamento_id")
                    user_update= request.POST.get("user_update")

                    equipamento_ob=get_object_or_404(equipamento,id=equipamento_id)
                    equipamento_ob.status=0
                    equipamento_ob.user_update=user_update
                    equipamento_ob.dateupdate=datetime.now()
                    equipamento_ob.save()
                                                                  
                    message='Equipamento eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
@csrf_exempt
def delete_equipamento_eleitoral(request):
  
  if request.method == "POST":
            try:
                    equipamento_id = request.POST.get("equipamento_id")
                    user_update= request.POST.get("user_update")

                    equipamento_ob=get_object_or_404(equipamento_eleitoral,id=equipamento_id)
                    equipamento_ob.status=0
                    equipamento_ob.user_update=user_update
                    equipamento_ob.dateupdate=datetime.now()
                    equipamento_ob.save()
                                                                  
                    message='Equipamento eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_equipamento_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         equipamento.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Equipamento eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
  
@csrf_exempt
def delete_equipamento_eleitoral_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         equipamento_eleitoral.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Equipamento eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def gestao_mobiliario(request):
   
    conselho_list = conselho.objects.all().filter(status=1)
    sala_list = sala.objects.all().filter(status=1)
    componente ='equipamento'

    query = '''
                    SELECT 
                    dm.id,
                    dm.descricao,
                    dm.data_entrada,
                    dm.serial_number,
                    dm.obs,
                    dm.tipo,
                    dm.provinencia,
                    dm.provinencia,
                    dm.carateristica,
                    kec.descricao as conselho,
                    IFNULL(ds.descricao,'') as sala
                    FROM departamentos_mobiliario dm
                    left join kit_eleitoral_conselho as kec on dm.conselho=kec.id
                    left join departamentos_sala as ds on dm.sala=ds.id
                    WHERE dm.STATUS=1 order by dm.id desc
                '''
    with connection.cursor() as cursor:
          cursor.execute(query)
          colunas = [col[0] for col in cursor.description] 
          mobiliario_list = [dict(zip(colunas, row)) for row in cursor.fetchall()]

          paginator = Paginator(mobiliario_list, 7)
          page_number = request.GET.get("page",1)  
          paginator_mobiliario = paginator.get_page(page_number)

          if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return render(request, 'mobiliario/index.html',{"mobiliario":paginator_mobiliario,"conselho":conselho_list,"sala":sala_list,"componente":componente})

    return render(request, 'mobiliario/index.html',{"mobiliario":paginator_mobiliario,"conselho":conselho_list,"sala":sala_list,"componente":componente})

def gestao_mobiliario_eleitoral(request):
   
    mobiliario_list = mobiliario_eleitoral.objects.all().filter(status=1)
    paginator = Paginator(mobiliario_list, 7)
    page_number = request.GET.get("page")  
    paginator_mobiliario = paginator.get_page(page_number)

    return render(request, 'Mobiliario_eleitoral/index.html',{"mobiliario":paginator_mobiliario})

@csrf_exempt
def add_mobiliario(request):
  
  if request.method == "POST":
            try:
                    descricao= request.POST.get("descricao")
                    data_entrada= request.POST.get("data_entrada_edit")
                    data_entrada= request.POST.get("data_entrada")
                    serial_number= request.POST.get("serial_number")
                    sala= request.POST.get("sala")
                    provinencia= request.POST.get("provinencia")
                    carateristica= request.POST.get("carateristica")
                    tipo= request.POST.get("tipo")
                    conselho= request.POST.get("conselho")
                    modelo= request.POST.get("modelo")
                    obs= request.POST.get("obs")
                    user_create= request.POST.get("user_create")

                    if conselho!="23":
                                    if descricao !="" and data_entrada !="" and provinencia!="" and conselho !="" and tipo !="" and carateristica !="":
                                      
                                         # validate=mobiliario.objects.filter(descricao=descricao).count()
                                         # if validate==0:

                                                    mobiliario.objects.create(
                                                                                descricao=descricao,
                                                                                data_entrada=data_entrada,
                                                                                serial_number=serial_number,
                                                                                conselho=conselho,
                                                                                carateristica=carateristica,
                                                                                provinencia=provinencia,
                                                                                modelo=modelo,
                                                                                tipo=tipo,
                                                                                obs=obs,
                                                                                user_create=user_create
                                                                                  )
                                                    message='Mobiliario registado com sucesso!!'
                                                    status= 'success'
                                                    return JsonResponse({'status':status, 'message': message })

                                         # else:
                                          
                                         #    message='Erro, este mobiliario já foi registado!!'
                                         #   status= 'error'
                                         #   return JsonResponse({'status':status, 'message': message })
                                    else:
                                      message='Erro, tem que preencher todos os campos obrigatorios!!'
                                      status= 'error'
                                      return JsonResponse({'status':status, 'message': message })
                    else:

                                if descricao !="" and data_entrada !="" and sala !="" and provinencia!="" and conselho !="" and tipo !="" and carateristica !="":
                                                
                                                   # validate=mobiliario.objects.filter(descricao=descricao).count()
                                                   #if validate==0:

                                                              mobiliario.objects.create(
                                                                                          descricao=descricao,
                                                                                          data_entrada=data_entrada,
                                                                                          serial_number=serial_number,
                                                                                          conselho=conselho,
                                                                                          carateristica=carateristica,
                                                                                          provinencia=provinencia,
                                                                                          modelo=modelo,
                                                                                          sala=sala,
                                                                                          tipo=tipo,
                                                                                          obs=obs,
                                                                                          user_create=user_create
                                                                                            )
                                                              message='Mobiliario registado com sucesso!!'
                                                              status= 'success'
                                                              return JsonResponse({'status':status, 'message': message })

                                                    #else:
                                                    
                                                    # message='Erro, este mobiliario já foi registado!!'
                                                    # status= 'error'
                                                    # return JsonResponse({'status':status, 'message': message })
                                else:
                                 message='Erro, tem que preencher todos os campos obrigatorio'
                                 status= 'error'
                                 return JsonResponse({'status':status, 'message': message })


                    

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_mobiliario_eleitoral(request):
  
  if request.method == "POST":
            try:
                    descricao= request.POST.get("descricao")
                    data_entrada= request.POST.get("data_entrada")
                    serial_number= request.POST.get("serial_number")
                    obs= request.POST.get("obs")
                    user_create= request.POST.get("user_create")

                    validate=mobiliario_eleitoral.objects.filter(descricao=descricao,serial_number=serial_number).count()

                    if descricao !="" and serial_number !="" and data_entrada !="":
                       
                          if validate==0:
                          

                                    mobiliario_eleitoral.objects.create(
                                                                data_entrada=data_entrada,
                                                                descricao=descricao,
                                                                serial_number=serial_number,
                                                                obs=obs,
                                                                user_create=user_create
                                                                  )
                                    message='Mobiliario registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                          else:
                            message='Erro, este mobiliario já foi registado!!'
                            status= 'error'
                            return JsonResponse({'status':status, 'message': message })
                    else:
                      message='Erro, tem que preencher todos os campos obrigatorios!!'
                      status= 'error'
                      return JsonResponse({'status':status, 'message': message })

                    

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_mobiliario(request):

    if request.method == "POST":
      try:
                      mobiliario_id = request.POST.get("mobiliario_id")

                      query = '''
                                      SELECT 
                                      dm.id,
                                      dm.descricao,
                                      dm.data_entrada,
                                      dm.serial_number,
                                      dm.obs,
                                      dm.tipo,
                                      dm.provinencia,
                                      carateristica,
                                      kec.descricao as conselho,
                                      kec.id as conselho_id,
                                      IFNULL(ds.descricao,'') as sala,
                                      ds.id as sala_id,
                                      dm.modelo
                                      FROM departamentos_mobiliario dm
                                      left join kit_eleitoral_conselho as kec on dm.conselho=kec.id
                                      left join departamentos_sala as ds on dm.sala=ds.id
                                      WHERE dm.STATUS=1 and dm.id=%s
                                  '''
                      with connection.cursor() as cursor:
                            cursor.execute(query,[mobiliario_id])
                            colunas = [col[0] for col in cursor.description] 
                            mobiliar = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                     
                      return JsonResponse({'resultado': mobiliar})

      except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_mobiliario_eleitoral(request):

    if request.method == "POST":
      try:
                      mobiliario_id = request.POST.get("mobiliario_id")
                      mobiliar= mobiliario_eleitoral.objects.filter(id=mobiliario_id)
                     
                      return JsonResponse(serialize("json", mobiliar),safe=False)

      except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def editar_mobiliario(request): 
  
  if request.method == "POST":
            try:
                    mobiliario_id = request.POST.get("mobiliario_id")
                    descricao= request.POST.get("descricao")
                    conselh= request.POST.get("conselho_edit")
                    sala_id= request.POST.get("sala_id")
                    obs= request.POST.get("obs")
                    user_update= request.POST.get("user_update")

                    if conselh!="23":

                                if conselh!="":
                                        mobiliario_ob=get_object_or_404(mobiliario,id=mobiliario_id)
                                        mobiliario_ob.descricao=descricao
                                        mobiliario_ob.obs=obs
                                        mobiliario_ob.sala=0
                                        mobiliario_ob.conselho=conselh
                                        mobiliario_ob.user_update=user_update
                                        mobiliario_ob.dateupdate=datetime.now()
                                        mobiliario_ob.save()
                                                                      
                                        message='Mobiliario alterado com sucesso!!'
                                        status= 'success'
                                        return JsonResponse({'status':status, 'message': message })
                                else:
                                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                                        status= 'error'
                                        return JsonResponse({'status':status, 'message': message })



                    else:
                               if conselh!="" and sala_id!="":
                                
                                    mobiliario_ob=get_object_or_404(mobiliario,id=mobiliario_id)
                                    mobiliario_ob.descricao=descricao
                                    mobiliario_ob.obs=obs
                                    mobiliario_ob.sala=sala_id
                                    mobiliario_ob.conselho=conselh
                                    mobiliario_ob.user_update=user_update
                                    mobiliario_ob.dateupdate=datetime.now()
                                    mobiliario_ob.save()
                                                                  
                                    message='Mobiliario alterado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })
                               else:
                                    message='Erro, tem que preencher todos os campos obrigatorios!!'
                                    status= 'error'
                                    return JsonResponse({'status':status, 'message': message })
                        

                  
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def editar_mobiliario_eleitoral(request):
  
  if request.method == "POST":
            try:
                    mobiliario_id = request.POST.get("mobiliario_id")
                    obs= request.POST.get("obs")
                    user_update= request.POST.get("user_update")


                    mobiliario_ob=get_object_or_404(mobiliario_eleitoral,id=mobiliario_id)
                    mobiliario_ob.obs=obs
                    mobiliario_ob.user_update=user_update
                    mobiliario_ob.dateupdate=datetime.now()
                    mobiliario_ob.save()
                                                                  
                    message='Mobiliario alterado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })

                  

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario(request):
  
  if request.method == "POST":
            try:
                    mobiliario_id = request.POST.get("mobiliario_id")
                    user_update= request.POST.get("user_update")

                    mobiliario_ob=get_object_or_404(mobiliario,id=mobiliario_id)
                    mobiliario_ob.status=0
                    mobiliario_ob.user_update=user_update
                    mobiliario_ob.dateupdate=datetime.now()
                    mobiliario_ob.save()
                                                                  
                    message='Mobiliario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_eleitoral(request):
  
  if request.method == "POST":
            try:
                    mobiliario_id = request.POST.get("mobiliario_id")
                    user_update= request.POST.get("user_update")

                    mobiliario_ob=get_object_or_404(mobiliario_eleitoral,id=mobiliario_id)
                    mobiliario_ob.status=0
                    mobiliario_ob.user_update=user_update
                    mobiliario_ob.dateupdate=datetime.now()
                    mobiliario_ob.save()
                                                                  
                    message='Mobiliario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_mb= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_mb:
                         id_mbs = id_mb.split(",") 
                         mobiliario.objects.filter(id__in=id_mbs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Mobiliario eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_eleitoral_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_mb= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_mb:
                         id_mbs = id_mb.split(",") 
                         mobiliario_eleitoral.objects.filter(id__in=id_mbs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Mobiliario eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_inventario_equipamento(request):
  
  if request.method == "POST":
            try:
                    provinencia= request.POST.get("provinencia")
                    equipamento= request.POST.get("equipamento")
                    estado= request.POST.get("estado")
                    user_create= request.POST.get("user_create")

                    if provinencia !="" and equipamento !="":

                                    inventario_equipamento.objects.create(
                                                                equipamento_id=equipamento,
                                                                provinencia=provinencia,
                                                                obs=estado,
                                                                user_create=user_create
                                                                  )
                                    message='Equipamento registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_inventario_equipamento_eleitoral(request):
  
  if request.method == "POST":
            try:
                    provinencia= request.POST.get("provinencia")
                    equipamento= request.POST.get("equipamento")
                    localizacao= request.POST.get("localizacao")
                    estado= request.POST.get("estado")
                    user_create= request.POST.get("user_create")

                    if provinencia !="" and equipamento !="" and localizacao !=""  and estado !="":

                                    inventario_equipamento_eleitoral.objects.create(
                                                                equipamento_id=equipamento,
                                                                localizacao=localizacao,
                                                                provinencia=provinencia,
                                                                obs=estado,
                                                                user_create=user_create
                                                                  )
                                    message='Equipamento registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_inventario_mobiliario(request):
  
  if request.method == "POST":
            try:
                    provinencia= request.POST.get("provinencia")
                    mobiliario= request.POST.get("mobiliario")
                    user_create= request.POST.get("user_create")
                    estado= request.POST.get("estado")

                    if provinencia !="" and equipamento !="":

                                    inventario_mobiliario.objects.create(
                                                                mobiliario_id=mobiliario,
                                                                provinencia=provinencia,
                                                                obs=estado,
                                                                user_create=user_create
                                                                  )
                                    message='Inventario mobiliario registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def add_inventario_mobiliario_eleitoral(request):
  
  if request.method == "POST":
            try:
                    provinencia= request.POST.get("provinencia")
                    mobiliario= request.POST.get("mobiliario")
                    localizacao= request.POST.get("localizacao")
                    estado= request.POST.get("estado")
                    user_create= request.POST.get("user_create")

                    if provinencia !="" and equipamento !="" and localizacao !=""  and estado !="":

                                    inventario_mobiliario_eleitoral.objects.create(
                                                                mobiliario_id=mobiliario,
                                                                localizacao=localizacao,
                                                                provinencia=provinencia,
                                                                obs=estado,
                                                                user_create=user_create
                                                                  )
                                    message='Inventario mobiliario registado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_equipamento_inventario(request):

    equipamento_id= request.POST.get("equipamento_id")

    print(equipamento_id)

    if request.method == "POST":

            try:
                  query = '''
                            SELECT 
                            die.id,
                            de.data_entrada,
                            kec.descricao as conselho,
                            die.obs,
                            die.provinencia,
                            de.descricao,
                            de.modelo,
                            de.marca,
                            de.mac_address,
                            de.serial_number,
                            de.id as equipamento_id
                            FROM 
                            departamentos_inventario_equipamento as die
                            inner join departamentos_equipamento as de on die.equipamento_id=de.id
                            inner join kit_eleitoral_conselho as kec on de.conselho=kec.id    
                            where die.id=%s
                    '''
                  with connection.cursor() as cursor:
                      cursor.execute(query,[equipamento_id])

                      colunas = [col[0] for col in cursor.description] 
                      resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                     
                  return JsonResponse({'resultado': resultados})

            except Exception as e:
                  return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_equipamento_eleitoral_inventario(request):

    equipamento_id= request.POST.get("equipamento_id")

    print(equipamento_id)

    if request.method == "POST":

            try:
                  query = '''
                            SELECT 
                            die.id,
                            de.data_entrada,
                            die.localizacao,
                            die.obs,
                            die.provinencia,
                            de.descricao,
                            de.modelo,
                            de.marca,
                            de.mac_address,
                            de.serial_number,
                            de.id as equipamento_id
                            FROM 
                            departamentos_inventario_equipamento_eleitoral as die
                            inner join departamentos_equipamento_eleitoral as de on die.equipamento_id=de.id
                            where die.id=%s
                    '''
                  with connection.cursor() as cursor:
                      cursor.execute(query,[equipamento_id])

                      colunas = [col[0] for col in cursor.description] 
                      resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                     
                  return JsonResponse({'resultado': resultados})

            except Exception as e:
                  return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
@csrf_exempt
def get_mobiliario_inventario(request):

    mobiliario_id= request.POST.get("mobiliario_id")

    if request.method == "POST":

            try:
                  query = '''
                            SELECT 
                            die.id,
                            de.data_entrada,
                            die.localizacao,
                            die.obs,
                            die.provinencia,
                            de.descricao,
                            de.id as mobiliario_id
                            FROM 
                            departamentos_inventario_mobiliario as die
                            inner join departamentos_mobiliario as de on die.mobiliario_id=de.id
                            where die.id=%s
                    '''
                  with connection.cursor() as cursor:
                      cursor.execute(query,[mobiliario_id])

                      colunas = [col[0] for col in cursor.description] 
                      resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                     
                  return JsonResponse({'resultado': resultados})

            except Exception as e:
                  return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def get_mobiliario_eleitoral_inventario(request):

    mobiliario_id= request.POST.get("mobiliario_id")

    if request.method == "POST":

            try:
                  query = '''
                            SELECT 
                            die.id,
                            de.data_entrada,
                            die.localizacao,
                            die.obs,
                            die.provinencia,
                            de.descricao,
                            de.id as mobiliario_id
                            FROM 
                            departamentos_inventario_mobiliario_eleitoral as die
                            inner join departamentos_mobiliario_eleitoral as de on die.mobiliario_id=de.id
                            where die.id=%s
                    '''
                  with connection.cursor() as cursor:
                      cursor.execute(query,[mobiliario_id])

                      colunas = [col[0] for col in cursor.description] 
                      resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                     
                  return JsonResponse({'resultado': resultados})

            except Exception as e:
                  return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def edit_inventario_equipamento(request):
  
  if request.method == "POST":
            try:
                                    id= request.POST.get("id")
                                    estado= request.POST.get("obs")
                                    user_update= request.POST.get("user_update")


                                    inventario_equipament=get_object_or_404(inventario_equipamento,id=id)
                                    inventario_equipament.obs=estado
                                    inventario_equipament.dateupdate=datetime.now()
                                    inventario_equipament.user_update=user_update
                                    inventario_equipament.save()
                                                                
                                    message='Inventario alterado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                  
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def edit_inventario_equipamente_eleitoral(request):
  
  if request.method == "POST":
            try:
                    id= request.POST.get("id")
                    localizacao= request.POST.get("localizacao")
                    estado= request.POST.get("obs")
                    user_update= request.POST.get("user_update")

                    if localizacao !=""  :

                                    inventario_equipament=get_object_or_404(inventario_equipamento_eleitoral,id=id)
                                    inventario_equipament.localizacao=localizacao
                                    inventario_equipament.obs=estado
                                    inventario_equipament.dateupdate=datetime.now()
                                    inventario_equipament.user_update=user_update
                                    inventario_equipament.save()
                                                                
                                    message='Inventario alterado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def edit_mobiliario_equipamento(request):
  
  if request.method == "POST":
            try:
                    id= request.POST.get("id")
                    localizacao= request.POST.get("localizacao")
                    estado= request.POST.get("obs")
                    user_update= request.POST.get("user_update")

                    if localizacao !="":

                                    inventario_mobiliari=get_object_or_404(inventario_mobiliario,id=id)
                                  
                                    inventario_mobiliari.localizacao=localizacao
                                    inventario_mobiliari.obs=estado
                                    inventario_mobiliari.dateupdate=datetime.now()
                                    inventario_mobiliari.user_update=user_update
                                    inventario_mobiliari.save()
                                                                
                                    message='Inventario mobiliario alterado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def edit_mobiliario_inventario_eleitoral(request):
  
  if request.method == "POST":
            try:
                    id= request.POST.get("id")
                    localizacao= request.POST.get("localizacao")
                    estado= request.POST.get("obs")
                    user_update= request.POST.get("user_update")

                    if localizacao !="":

                                    inventario_mobiliari=get_object_or_404(inventario_mobiliario_eleitoral,id=id)
                                    inventario_mobiliari.localizacao=localizacao
                                    inventario_mobiliari.obs=estado
                                    inventario_mobiliari.dateupdate=datetime.now()
                                    inventario_mobiliari.user_update=user_update
                                    inventario_mobiliari.save()
                                                                
                                    message='Inventario mobiliario alterado com sucesso!!'
                                    status= 'success'
                                    return JsonResponse({'status':status, 'message': message })

                    else:
                        message='Erro, tem que preencher todos os campos obrigatorios!!'
                        status= 'error'
                        return JsonResponse({'status':status, 'message': message })
            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_equipamento_inventario(request):
  
  if request.method == "POST":
            try:
                    inventario_equipamento_id = request.POST.get("inventario_equipamento_id")
                    user_update= request.POST.get("user_update")

                    inventario_equipament=get_object_or_404(inventario_equipamento,id=inventario_equipamento_id)
                    inventario_equipament.status=0
                    inventario_equipament.user_update=user_update
                    inventario_equipament.dateupdate=datetime.now()
                    inventario_equipament.save()
                                                                  
                    message='Inventario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
@csrf_exempt
def delete_equipamento_eleitoral_inventario(request):
  
  if request.method == "POST":
            try:
                    inventario_equipamento_id = request.POST.get("inventario_equipamento_id")
                    user_update= request.POST.get("user_update")

                    inventario_equipament=get_object_or_404(inventario_equipamento_eleitoral,id=inventario_equipamento_id)
                    inventario_equipament.status=0
                    inventario_equipament.user_update=user_update
                    inventario_equipament.dateupdate=datetime.now()
                    inventario_equipament.save()
                                                                  
                    message='Inventario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
@csrf_exempt
def delete_mobiliario_inventario(request):
  
  if request.method == "POST":
            try:
                    inventario_mobiliario_id = request.POST.get("inventario_mobiliario_id")
                    user_update= request.POST.get("user_update")

                    inventario_mobiliari=get_object_or_404(inventario_mobiliario,id=inventario_mobiliario_id)
                    inventario_mobiliari.status=0
                    inventario_mobiliari.user_update=user_update
                    inventario_mobiliari.dateupdate=datetime.now()
                    inventario_mobiliari.save()
                                                                  
                    message='Inventario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_eleitoral_inventario(request):
  
  if request.method == "POST":
            try:
                    inventario_mobiliario_id = request.POST.get("inventario_mobiliario_id")
                    user_update= request.POST.get("user_update")

                    inventario_mobiliari=get_object_or_404(inventario_mobiliario_eleitoral,id=inventario_mobiliario_id)
                    inventario_mobiliari.status=0
                    inventario_mobiliari.user_update=user_update
                    inventario_mobiliari.dateupdate=datetime.now()
                    inventario_mobiliari.save()
                                                                  
                    message='Inventario eliminado com sucesso!!'
                    status= 'success'
                    return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def delete_equipamento_inventario_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         inventario_equipamento.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Equipamento eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_equipamento_inventario_eleitoral_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         inventario_equipamento_eleitoral.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Equipamento eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_inventario_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         inventario_mobiliario.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Inventario eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@csrf_exempt
def delete_mobiliario_inventario_eleitoral_checkbox(request):
  
  if request.method == "POST":
            try:
                   id_eq= request.POST.get("id")
                   id_user= request.POST.get("id_user")

                   if id_eq:
                         id_eqs = id_eq.split(",") 
                         inventario_mobiliario_eleitoral.objects.filter(id__in=id_eqs).update(status=0,user_update=id_user,dateupdate=datetime.now())
                                                                  
                   message='Inventario eliminado com sucesso!!'
                   status= 'success'
                   return JsonResponse({'status':status, 'message': message })
         

            except Exception as e:
             return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
  
  


def exportar_inventario_equipamento_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Inventario_Equipamento'

    query = '''
                      SELECT 
                      die.id as inventario_equipamento_id,
                      de.data_entrada,
                      die.localizacao,
                      die.obs,
                      die.provinencia,
                      de.descricao,
                      de.modelo,
                      de.marca,
                      de.mac_address,
                      de.serial_number
                      FROM 
                      departamentos_inventario_equipamento as die
                      inner join departamentos_equipamento as de on die.equipamento_id=de.id
                      where die.status=1
                  '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Equipamento', 'Localização','Modelo','Serial Number','Provinçia','Mac Addres','Marca','Obs'])

                    # Dados
                    for equipamento in resultados:
                        folha.append([equipamento['inventario_equipamento_id'], equipamento['data_entrada'], equipamento['descricao'],equipamento['localizacao'],equipamento['modelo'],equipamento['serial_number'],equipamento['provinencia'],equipamento['mac_address'],equipamento['marca'],equipamento['obs']])

                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Inventario.xlsx'
                    workbook.save(response)

    return response

def exportar_inventario_mobiliario_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Inventario_Equipamento'

    query = '''
                    SELECT 
                    die.id,
                    die.data_entrada,
                    de.localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao
                    FROM 
                    departamentos_inventario_mobiliario as die
                    inner join departamentos_mobiliario as de on die.mobiliario_id=de.id
                    where die.status=1
                  '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Mobiliario', 'Localização','Provinçia','OBS'])

                    # Dados
                    for mobiliario in resultados:
                        folha.append([mobiliario['id'], mobiliario['data_entrada'], mobiliario['descricao'],mobiliario['localizacao'],mobiliario['provinencia'],mobiliario['obs']])

                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Inventario.xlsx'
                    workbook.save(response)

    return response

def exportar_inventario_mobiliario_eleitoral_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Inventario_Mobiliario_Eleitoral'

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    die.localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao
                    FROM 
                    departamentos_inventario_mobiliario_eleitoral as die
                    inner join departamentos_mobiliario_eleitoral as de on die.mobiliario_id=de.id
                    where die.status=1
                  '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Mobiliario', 'Localização','Provinçia','OBS'])

                    # Dados
                    for mobiliario in resultados:
                        folha.append([mobiliario['id'], mobiliario['data_entrada'], mobiliario['descricao'],mobiliario['localizacao'],mobiliario['provinencia'],mobiliario['obs']])

                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Inventario_Mobiliario_Eleitoral.xlsx'
                    workbook.save(response)

    return response

def exportar_inventario_equipamento_eleitoral_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Inventario_Equipamento_eleitoral'

    query = '''
                    SELECT 
                    die.id,
                    de.data_entrada,
                    die.localizacao,
                    die.obs,
                    die.provinencia,
                    de.descricao,
                    de.modelo,
                    de.marca,
                    de.mac_address,
                    de.serial_number
                    FROM 
                    departamentos_inventario_equipamento_eleitoral as die
                    inner join departamentos_equipamento_eleitoral as de on die.equipamento_id=de.id
                    where die.status=1
                  '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Equipamento', 'Localização','Modelo','Serial Number','Provinçia','Mac Addres','Marca','Obs'])

                    # Dados
                    for equipamento in resultados:
                        folha.append([equipamento['id'], equipamento['data_entrada'], equipamento['descricao'],equipamento['localizacao'],equipamento['modelo'],equipamento['serial_number'],equipamento['provinencia'],equipamento['mac_address'],equipamento['marca'],equipamento['obs']])

                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Inventario_Equipamento_eleitoral.xlsx'
                    workbook.save(response)

    return response
  
def exportar_equipamento_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Equipamento'

    query = '''
                SELECT 
                departamentos_equipamento.id,
                departamentos_equipamento.descricao,
                mac_address,
                data_entrada,
                provinencia,
                marca,
                modelo,
                obs,
                serial_number,
                tipo,
                IFNULL(departamentos_sala.descricao,'') as sala,
                kit_eleitoral_conselho.descricao as conselho
                FROM 
                departamentos_equipamento
                left join departamentos_sala on departamentos_equipamento.sala=departamentos_sala.id
                left join kit_eleitoral_conselho on departamentos_equipamento.conselho=kit_eleitoral_conselho.id
                where departamentos_equipamento.status=1
              '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Equipamento','provinencia','Localização','Sala','Modelo','Serial Number','Mac Addres','Marca','Tipo Item','Obs'])
                    folha.font = Font(bold=True)

                    # Dados
                    for equipamento in resultados:
                        folha.append([equipamento['id'], equipamento['data_entrada'], equipamento['descricao'],equipamento['provinencia'],equipamento['conselho'],equipamento['sala'],equipamento['modelo'],equipamento['serial_number'],equipamento['mac_address'],equipamento['marca'],equipamento['tipo'],equipamento['obs']])
                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Equipamento_Excel.xlsx'
                    workbook.save(response)

    return response

def exportar_Mobiliario_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Mobiliario'

    query = '''
                SELECT 
                dm.id,
                dm.descricao,
                dm.data_entrada,
                dm.serial_number,
                dm.obs,
                dm.tipo,
                dm.provinencia,
                dm.provinencia,
                dm.carateristica,
                kec.descricao as conselho,
                IFNULL(ds.descricao,'') as sala
                FROM departamentos_mobiliario dm
                left join kit_eleitoral_conselho as kec on dm.conselho=kec.id
                left join departamentos_sala as ds on dm.sala=ds.id
                WHERE dm.STATUS=1
              '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Mobiliario','provinencia','Localização','Sala','Serial Number','carateristica','Tipo Item','Obs'])

                    # Dados
                    for equipamento in resultados:
                        folha.append([equipamento['id'], equipamento['data_entrada'], equipamento['descricao'],equipamento['provinencia'],equipamento['conselho'],equipamento['sala'],equipamento['serial_number'],equipamento['carateristica'],equipamento['tipo'],equipamento['obs']])
                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Mobiliario_Excel.xlsx'
                    workbook.save(response)

    return response

def exportar_kit_excel(request):
    # Criar workbook e folha
    workbook = openpyxl.Workbook()
    folha = workbook.active
    folha.title = 'Mobiliario'

    query = '''
                SELECT 
                dm.id,
                dm.descricao,
                dm.data_entrada,
                dm.serial_number,
                dm.obs,
                dm.tipo,
                dm.provinencia,
                dm.provinencia,
                dm.carateristica,
                kec.descricao as conselho,
                IFNULL(ds.descricao,'') as sala
                FROM departamentos_mobiliario dm
                left join kit_eleitoral_conselho as kec on dm.conselho=kec.id
                left join departamentos_sala as ds on dm.sala=ds.id
                WHERE dm.STATUS=1
              '''
    with connection.cursor() as cursor:
                    cursor.execute(query)

                    colunas = [col[0] for col in cursor.description] 
                    resultados = [dict(zip(colunas, row)) for row in cursor.fetchall()]
                    print(resultados)

                  # Cabeçalhos
                    folha.append(['id','Data Entrada', 'Mobiliario','provinencia','Localização','Sala','Serial Number','carateristica','Tipo Item','Obs'])

                    # Dados
                    for equipamento in resultados:
                        folha.append([equipamento['id'], equipamento['data_entrada'], equipamento['descricao'],equipamento['provinencia'],equipamento['conselho'],equipamento['sala'],equipamento['serial_number'],equipamento['carateristica'],equipamento['tipo'],equipamento['obs']])
                    # Preparar resposta HTTP
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename=Mobiliario_Excel.xlsx'
                    workbook.save(response)

    return response